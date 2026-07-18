"""Outreach contacts: published channels allowed only inside `contacts`,
never inferred elsewhere; Excel Respondents sheet add/create; manual-only flow.
Run with: python -m unittest tests.test_respondent_contacts"""
import json
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from src import export_excel as xl
from src import onepager, respondents, runs

TODAY = date.today().isoformat()


def _cand(**over):
    c = {"name": "Иван Иванов", "role": "CIO", "org": "Банк N",
         "why_relevant": "внедрял BPM", "addresses": ["buying_behavior"],
         "priority": 1, "profile_url": "https://conf.ru/speakers/ivanov",
         "sources": ["https://conf.ru/speakers/ivanov"],
         "confidence": "high", "role_current": True, "verified_on": TODAY}
    c.update(over)
    return c


def _codes(doc, scope="market", entity=""):
    return {(i["code"], i["severity"])
            for i in respondents.validate_respondents(doc, set(), scope, entity)}


class TestContactChannels(unittest.TestCase):
    def test_published_channels_in_contacts_accepted(self):
        doc = {"scope": "market", "candidates": [_cand(contacts={
            "telegram": "@ivanov_cio", "email": "i.ivanov@bank-n.ru",
            "linkedin": "https://linkedin.com/in/ivanov",
            "phone": "+7 495 123-45-67"})]}
        rejects = [i for i in respondents.validate_respondents(doc, set(), "market")
                   if i["severity"] == "reject"]
        self.assertEqual(rejects, [])

    def test_contact_outside_contacts_still_rejected(self):
        # the whole point of the old net is preserved for free-text fields
        self.assertIn(("private-contact", "reject"),
                      _codes({"scope": "market",
                              "candidates": [_cand(why_relevant="пишите i@bank.ru")]}))
        self.assertIn(("private-contact", "reject"),
                      _codes({"scope": "market",
                              "candidates": [_cand(email="i@bank.ru")]}))  # top-level

    def test_malformed_channels_rejected(self):
        for ct, bad in (({"email": "not-an-email"}, "email"),
                        ({"telegram": "ivanov"}, "telegram"),   # missing @ / t.me
                        ({"linkedin": "https://vk.com/x"}, "linkedin"),
                        ({"phone": "звоните завтра"}, "phone")):
            codes = _codes({"scope": "market", "candidates": [_cand(contacts=ct)]})
            self.assertIn(("bad-contact", "reject"), codes, ct)

    def test_telegram_channel_and_handle_forms_accepted(self):
        for tg in ("@ivanov_cio", "https://t.me/ivanov_cio", "https://t.me/s/true_conf"):
            codes = _codes({"scope": "market",
                            "candidates": [_cand(contacts={"telegram": tg})]})
            self.assertNotIn(("bad-contact", "reject"), codes, tg)

    def test_channel_without_source_rejected(self):
        doc = {"scope": "market",
               "candidates": [_cand(sources=[], contacts={"telegram": "@x_cio"})]}
        # sources also required-empty, but the channel-needs-a-source rule fires
        self.assertIn(("unsourced", "reject"), _codes(doc))

    def test_no_channel_is_warn_not_reject(self):
        codes = _codes({"scope": "market", "candidates": [_cand()]})
        self.assertIn(("no-channel", "warn"), codes)
        self.assertNotIn(("no-channel", "reject"), codes)

    def test_autofix_migrates_top_level_channel_into_contacts(self):
        doc = {"scope": "market", "candidates": [_cand(telegram="@ivanov_cio")]}
        notes = respondents.autofix_doc(doc)
        c = doc["candidates"][0]
        self.assertNotIn("telegram", set(c) - {"contacts"})
        self.assertEqual(c["contacts"]["telegram"], "@ivanov_cio")
        self.assertTrue(any("contacts.telegram" in n for n in notes))


class TestExcelSheet(unittest.TestCase):
    def _rows(self):
        return [{"target": "Market", "name": "Иван", "telegram": "@x",
                 "email": "i@b.ru", "priority": 1}]

    def test_creates_workbook_with_respondents_first(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "new.xlsx"
            xl.write_respondents_sheet(out, respondents.RESP_COLUMNS, self._rows())
            from openpyxl import load_workbook
            wb = load_workbook(out)
            self.assertEqual(wb.sheetnames[0], "Respondents")
            self.assertEqual([c.value for c in wb["Respondents"][1]][:2],
                             ["target", "name"])

    def test_adds_sheet_to_existing_quant_workbook(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "deliverable.xlsx"
            from openpyxl import Workbook, load_workbook
            wb = Workbook(); wb.active.title = "Research"
            wb.active["A1"] = "brand_name"; wb.save(out)
            xl.write_respondents_sheet(out, respondents.RESP_COLUMNS, self._rows())
            wb2 = load_workbook(out)
            self.assertIn("Research", wb2.sheetnames)      # quant sheet preserved
            self.assertIn("Respondents", wb2.sheetnames)

    def test_stale_respondents_sheet_replaced_not_duplicated(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "d.xlsx"
            xl.write_respondents_sheet(out, respondents.RESP_COLUMNS, self._rows())
            xl.write_respondents_sheet(out, respondents.RESP_COLUMNS,
                                       self._rows() + self._rows())
            from openpyxl import load_workbook
            wb = load_workbook(out)
            self.assertEqual(wb.sheetnames.count("Respondents"), 1)
            self.assertEqual(wb["Respondents"].max_row, 3)   # header + 2 rows


class TestManualOnlyDeliverable(unittest.TestCase):
    def test_manual_only_builds_standalone_respondents_workbook(self):
        with tempfile.TemporaryDirectory() as td:
            with patch.object(runs, "LOGS", Path(td)):
                rd = onepager.create_manual_run("пилоты BPM")
                onepager.setup(rd, "выбрать вендора", {"Новая Ко": "partner"},
                               manual={"Новая Ко": {"segment": "BPM", "notes": "n"}},
                               require_goal=False)
                respondents.resp_path(rd, "новая-ко").write_text(json.dumps(
                    {"scope": "company", "entity": "Новая Ко",
                     "candidates": [_cand(name="П П", org="Новая Ко",
                                          contacts={"telegram": "@pp_bpm"})]},
                    ensure_ascii=False), encoding="utf-8")
                respondents.resp_path(rd, respondents.MARKET_STEM).write_text(
                    json.dumps({"scope": "market", "candidates": []},
                               ensure_ascii=False), encoding="utf-8")
                out = respondents.build_contacts_xlsx(rd)
                self.assertIsNotNone(out)
                from openpyxl import load_workbook
                wb = load_workbook(out)
                self.assertEqual(wb.sheetnames, ["Respondents"])   # no quant sheet
                vals = [c.value for c in wb["Respondents"][2]]
                self.assertIn("'@pp_bpm", vals)                    # channel present (formula-guarded)
                self.assertIn("Новая Ко", vals)                    # manual company


if __name__ == "__main__":
    unittest.main()


class TestContactGrounding(unittest.TestCase):
    """Audit fix: a channel must appear in a FETCHED page, not merely sit next
    to a cited URL. DeepSeek app-tools path."""
    from src.web_tools import SourceLog as _SL

    URL = "https://conf.ru/speakers/ivanov"   # == _cand() profile_url & sources

    def _log(self, text):
        from src.web_tools import SourceLog
        log = SourceLog()
        log.log_fetch(self.URL, {"url": self.URL, "final_url": self.URL,
                                 "title": "t", "text": text})
        return log

    def test_fabricated_email_stripped_even_with_real_source(self):
        log = self._log("Иван Иванов, CIO Банк N — спикер")   # no email on page
        doc = {"scope": "market",
               "candidates": [_cand(contacts={"email": "ivan.fake@bank-n.ru"})]}
        details = respondents.ground_candidates(doc, log)
        self.assertTrue(any("ungrounded email" in d for d in details))
        self.assertNotIn("email", doc["candidates"][0].get("contacts", {}))

    def test_published_contacts_survive_grounding(self):
        log = self._log("Иван Иванов — email ivan.real@bank-n.ru, тел +7 916 555-44-33, @ivanov_cio")
        doc = {"scope": "market", "candidates": [_cand(contacts={
            "email": "ivan.real@bank-n.ru", "phone": "+7 916 555-44-33",
            "telegram": "@ivanov_cio"})]}
        details = respondents.ground_candidates(doc, log)
        self.assertEqual([d for d in details if "ungrounded" in d and "contact" not in d], [])
        self.assertEqual(set(doc["candidates"][0]["contacts"]),
                         {"email", "phone", "telegram"})

    def test_repair_reaudits_changed_contact_same_profile(self):
        prev = {"scope": "market",
                "candidates": [_cand(contacts={"email": "real@bank-n.ru"})]}
        new = {"scope": "market",
               "candidates": [_cand(contacts={"email": "invented@bank-n.ru"})]}
        # repair opened nothing; changed email must NOT be trusted
        details = respondents.ground_candidates(new, self._log(""), prev_doc=prev)
        self.assertTrue(any("ungrounded email" in d for d in details))
        self.assertNotIn("email", new["candidates"][0].get("contacts", {}))

    def test_repair_trusts_unchanged_contact(self):
        prev = {"scope": "market",
                "candidates": [_cand(contacts={"email": "real@bank-n.ru"})]}
        new = {"scope": "market",
               "candidates": [_cand(contacts={"email": "real@bank-n.ru"})]}
        details = respondents.ground_candidates(new, self._log(""), prev_doc=prev)
        self.assertEqual(new["candidates"][0]["contacts"]["email"], "real@bank-n.ru")
        self.assertEqual([d for d in details if "email" in d], [])


class TestGenericInbox(unittest.TestCase):
    def test_generic_mailbox_rejected(self):
        for addr in ("info@bank-n.ru", "pr@bank-n.ru", "sales@bank-n.ru"):
            codes = _codes({"scope": "market",
                            "candidates": [_cand(contacts={"email": addr})]})
            self.assertIn(("bad-contact", "reject"), codes, addr)

    def test_personal_mailbox_accepted(self):
        codes = _codes({"scope": "market",
                        "candidates": [_cand(contacts={"email": "i.ivanov@bank-n.ru"})]})
        self.assertNotIn(("bad-contact", "reject"), codes)


class TestExcelFormulaInjection(unittest.TestCase):
    def test_leading_formula_chars_neutralized(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "d.xlsx"
            xl.write_respondents_sheet(out, respondents.RESP_COLUMNS, [
                {"name": '=HYPERLINK("http://evil")', "role": "+1", "org": "@x"}])
            from openpyxl import load_workbook
            ws = load_workbook(out)["Respondents"]
            for col, txt in (("B", "=HYPER"), ("C", "+1"), ("D", "@x")):
                cell = ws[f"{col}2"]
                self.assertEqual(cell.data_type, "s", col)      # string, not formula
                self.assertTrue(str(cell.value).startswith("'"), col)
