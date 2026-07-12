"""Deterministic autofix, computed-field sourcing, quota fallback, Excel
highlight. Run with: python -m unittest tests.test_autofix_and_quota"""
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from src import api_runner, export_excel, gate, runs, web_tools


def _f(v, src="https://x.ru/a"):
    return {"value": v, "source": src}


class TestComputedFieldSourcing(unittest.TestCase):
    def _codes(self, fields):
        return {(i["field"], i["code"])
                for i in gate.validate_record({"entity": "Т", "fields": fields},
                                              None, None)}

    def test_projection_without_url_ok_when_inputs_sourced(self):
        codes = self._codes({"total_revenue_2024": _f("100 млн ₽"),
                             "total_revenue_2025": _f("120 млн ₽"),
                             "revenue_2026_projection": {"value": "~140 млн ₽"}})
        self.assertNotIn(("revenue_2026_projection", "unsourced"), codes)

    def test_projection_still_unsourced_when_inputs_unsourced(self):
        codes = self._codes({"total_revenue_2025": {"value": "120 млн ₽"},
                             "revenue_2026_projection": {"value": "~140 млн ₽"}})
        self.assertIn(("revenue_2026_projection", "unsourced"), codes)

    def test_yoy_without_url_ok_when_pair_sourced(self):
        codes = self._codes({"ebitda_2024": _f("10 млн ₽"),
                             "ebitda_2025": _f("12 млн ₽"),
                             "ebitda_yoy_24_25": {"value": "20%"}})
        self.assertNotIn(("ebitda_yoy_24_25", "unsourced"), codes)

    def test_long_digitfree_prose_with_stub_word_not_placeholder(self):
        # СберТех case: 400+ chars of real business_model text, no digits,
        # contains «не раскры…»
        v = ("Модель: лицензирование и подписка на продукты платформы. "
             "Выручка по продуктам не раскрывается. " + "Продажи прямые. " * 25)
        self.assertGreater(len(v), 200)
        self.assertFalse(gate.is_placeholder(v))


class TestAutofix(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.rd = Path(self.tmp.name)
        (self.rd / "agent_runs").mkdir()
        (self.rd / "companies.json").write_text(json.dumps({
            "market": "m", "segments": ["Low-code BPM платформы", "Банковский BPM"],
            "companies": [{"brand": "Тест", "segment": "Low-code BPM платформы"}]},
            ensure_ascii=False), encoding="utf-8")
        rec = {"entity": "Тест", "entity_match": {"entity_type": "company_product"},
               "fields": {
                   "segment": _f("Low-code BPM-платформы"),
                   "inn": _f("7715560268 (ООО); 9715302870 (ПАО)"),
                   "total_revenue_2024": _f("100 млн ₽"),
                   "total_revenue_2025": _f("150 млн ₽"),
                   "description": _f("к" * 130), "key_products": _f("Тест"),
                   "business_model": _f("подписка"), "target_customers": _f("МСБ"),
                   "positioning": _f("лидер"),
               }}
        (self.rd / "agent_runs" / "тест_record.json").write_text(
            json.dumps(rec, ensure_ascii=False), encoding="utf-8")
        (self.rd / "run.json").write_text(json.dumps(
            {"run_id": "t", "market": "m", "depth": "superficial",
             "model": "chatgpt", "output_language": "Russian", "status": "x"},
            ensure_ascii=False), encoding="utf-8")

    def tearDown(self):
        self.tmp.cleanup()

    def test_deterministic_fixes_without_web(self):
        fixed = runs.autofix_records(self.rd)
        self.assertIn("Тест", fixed)
        rec = json.loads((self.rd / "agent_runs" / "тест_record.json").read_text())
        f = rec["fields"]
        self.assertEqual(f["segment"]["value"], "Low-code BPM платформы")  # snap
        self.assertEqual(rec["entity_match"]["entity_type"], "product")   # map
        self.assertEqual(f["inn"]["value"], "7715560268")                 # 1st valid
        self.assertEqual(f["revenue_yoy_24_25"]["value"], "50%")          # computed

    def test_accepted_records_untouched(self):
        # make the record fully acceptable first
        runs.autofix_records(self.rd)
        p = self.rd / "agent_runs" / "тест_record.json"
        before = p.read_text()
        g = runs.run_gate(self.rd, write_report=False)
        if g["accepted"]:                       # if it passes, autofix must no-op
            runs.autofix_records(self.rd)
            self.assertEqual(p.read_text(), before)


class TestQuotaFallback(unittest.TestCase):
    def tearDown(self):
        web_tools.reset_quota_flag()

    def test_web_search_short_circuits_when_flag_set(self):
        web_tools.QUOTA_EXHAUSTED = True
        with patch("src.web_tools.requests.get") as get:
            with self.assertRaises(web_tools.SearchQuotaExhausted):
                web_tools.web_search("q")
            get.assert_not_called()             # no HTTP at all

    def test_brave_402_sets_sticky_flag(self):
        resp = unittest.mock.MagicMock(status_code=402)
        with patch("src.web_tools.requests.get", return_value=resp), \
             patch.dict("os.environ", {"SEARCH_API_KEY": "k"}):
            with self.assertRaises(web_tools.SearchQuotaExhausted):
                web_tools.web_search("q")
        self.assertTrue(web_tools.QUOTA_EXHAUSTED)


class TestExcelUnresolvedHighlight(unittest.TestCase):
    def test_unresolved_cells_empty_and_yellow(self):
        rec = {"entity": "Тест", "fields": {
                   "headcount": {"value": "", "source": ""},
                   "description": _f("описание компании")},
               "review_flags": ["unresolved: headcount — нет поисковой квоты"]}
        cols = export_excel.get_columns()
        row = export_excel._record_to_row(rec, cols)
        self.assertIn("headcount", row["_unresolved"])
        self.assertEqual(row["headcount"], "")            # empty, no marker text
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "t.xlsx"
            export_excel.write_xlsx(cols, [row], out)
            from openpyxl import load_workbook
            ws = load_workbook(out).active
            ci = cols.index("headcount") + 1
            cell = ws.cell(row=2, column=ci)
            self.assertEqual(cell.value, None)             # empty cell
            self.assertIn("FFF599", str(cell.fill.fgColor.rgb))   # yellow
            self.assertNotIn("MANUAL", str(cell.value))


class TestRepairLivelockGuard(unittest.TestCase):
    def test_blank_unresolved_skips_required_fields(self):
        rec = {"fields": {"headcount": _f("120"), "description": _f("текст")}}
        issues = [{"field": "headcount", "code": "unsourced", "severity": "reject"},
                  {"field": "description", "code": "unsourced", "severity": "reject"}]
        blanked = api_runner._blank_unresolved(rec, issues, "тест")
        self.assertEqual(blanked, ["headcount"])           # required field kept
        self.assertEqual(rec["fields"]["headcount"]["value"], "")
        self.assertEqual(rec["fields"]["description"]["value"], "текст")
        self.assertEqual(rec["review_flags"], ["unresolved: headcount — тест"])


if __name__ == "__main__":
    unittest.main()
