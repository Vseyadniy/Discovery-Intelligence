"""
Export the research table as an easy-to-read Excel workbook (and an Airtable-ready
CSV). The deliverable for a run is a **table, one row per company**, not a Markdown
report.

Three sources:
  --from-records outputs/agent_runs → the verifier *record* JSONs produced by an
                                     emulated run (render_prompts flow). The table
                                     is a deterministic projection of those records,
                                     not a hand-typed CSV. Also writes the CSV.
  --csv  outputs/gpu_iaas_ru.csv   → a filled table taken as-is (columns verbatim).
  --from-kb                        → the consolidated_record view in db/kb.sqlite
                                     (a real local pipeline run), pivoted to one
                                     row per entity.

Usage:
  python -m src.export_excel --from-records outputs/agent_runs --out docs/test_gpu/GPU_IaaS_RU.xlsx
  python -m src.export_excel --csv outputs/gpu_iaas_ru.csv --out docs/test_gpu/GPU_IaaS_RU.xlsx
  python -m src.export_excel --from-kb --out docs/test_gpu/GPU_IaaS_RU.xlsx
  python -m src.export_excel --template outputs/gpu_iaas_ru.csv   # write a blank GPU template

The CSV that feeds (or is written next to) the workbook imports directly into
Airtable as a new table. openpyxl is the only extra dependency.
"""
from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

# The deliverable columns, in reading order. Kept in sync with config/schema.yaml
# plus the provenance/notes columns a client review needs.
#
# Financials are ordered YEAR BY YEAR and carry a two-level Excel outline:
# level 1 = the whole financial block, level 2 = each year's columns — so the
# workbook drills down: [+ финансы] → [+2022] [+2023] [+2024] [+2025] [+2026].
FINANCIAL_YEAR_GROUPS = {
    2022: ["product_revenue_2022", "total_revenue_2022", "ebitda_2022"],
    2023: ["product_revenue_2023", "total_revenue_2023", "ebitda_2023"],
    2024: ["product_revenue_2024", "total_revenue_2024", "ebitda_2024"],
    2025: ["product_revenue_2025", "total_revenue_2025", "ebitda_2025",
           "revenue_yoy_24_25", "product_rev_yoy_24_25", "ebitda_yoy_24_25"],
    2026: ["revenue_2026_projection"],
}
_YEAR_COLS = [c for y in sorted(FINANCIAL_YEAR_GROUPS) for c in FINANCIAL_YEAR_GROUPS[y]]
FINANCIAL_COLUMNS = _YEAR_COLS + ["product_revenue_source"]

_PERCENT_COLS = {"revenue_yoy_24_25", "product_rev_yoy_24_25", "ebitda_yoy_24_25"}

COLUMNS = [
    "brand_name", "legal_entity_name", "inn", "website", "entity_type",
    "segment", "description", "business_model", "target_customers", "positioning",
    *FINANCIAL_COLUMNS,
    "headcount", "key_products", "other_products", "latest_news", "confidence_entity_match",
    "data_quality", "financial_sources", "other_sources",
]

# Fields whose sources land in `financial_sources` (registry + money + size);
# everything else goes to `other_sources` (company, product, business).
_FINANCIAL_SOURCE_FIELDS = set(FINANCIAL_COLUMNS) | {
    "legal_entity_name", "inn", "headcount", "revenue",
}


def custom_columns() -> list[str]:
    """App-added research columns (config/custom_fields.yaml), if any."""
    import yaml
    f = ROOT / "config" / "custom_fields.yaml"
    if not f.exists():
        return []
    extra = (yaml.safe_load(f.read_text(encoding="utf-8")) or {}).get("fields") or []
    return [x["name"] for x in extra if x.get("name") and x["name"] not in COLUMNS]


def get_columns() -> list[str]:
    """The deliverable columns including custom ones (inserted before latest_news)."""
    cols = list(COLUMNS)
    cc = custom_columns()
    if cc:
        i = cols.index("latest_news")
        cols[i:i] = cc
    return cols

INPUT_CSV = ROOT / "inputs" / "input_entities.csv"

# Per-column display width (Excel units), capped for readability.
_WIDTHS = {
    "brand_name": 18, "legal_entity_name": 30, "inn": 14, "website": 24,
    "entity_type": 13, "segment": 22, "description": 56,
    "business_model": 32, "target_customers": 32, "positioning": 32,
    **{c: 15 for c in FINANCIAL_COLUMNS},
    **{c: 13 for c in _PERCENT_COLS},
    "revenue_2026_projection": 16, "product_revenue_source": 40,
    "headcount": 10, "key_products": 40, "other_products": 36, "latest_news": 48,
    "confidence_entity_match": 14, "data_quality": 20,
    "financial_sources": 45, "other_sources": 45, "sources": 50, "notes": 40,
}


def _seed_rows(subsector: str | None = None) -> list[dict]:
    """Blank rows seeded from inputs/input_entities.csv (any market), research fields empty."""
    rows = []
    want = (subsector or "").strip().lower()
    with INPUT_CSV.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            if want and (r.get("subsector") or "").strip().lower() != want:
                continue
            rows.append({
                "brand_name": r.get("seed_name", ""),
                "inn": r.get("inn", ""),
                "website": r.get("website", ""),
                "segment": r.get("subsector", ""),
            })
    return rows


def write_template(path: Path, subsector: str | None = None) -> None:
    """Write a blank research CSV seeded from the input list for the researcher to fill."""
    seeds = _seed_rows(subsector)
    cols = get_columns()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for seed in seeds:
            w.writerow({c: seed.get(c, "") for c in cols})
    scope = f" ({subsector})" if subsector else ""
    print(f"[template] wrote {len(seeds)}-row blank table{scope} → {path}")


def read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with path.open(encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        headers = reader.fieldnames or COLUMNS
    return list(headers), rows


def read_kb(db_path: Path) -> tuple[list[str], list[dict]]:
    """Pivot the consolidated_record view to one row per entity."""
    if not db_path.exists():
        raise SystemExit(f"No KB at {db_path}. Run the pipeline first, or use --csv.")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.execute("SELECT * FROM consolidated_record")
        recs = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()
    # consolidated_record is long (entity, field, value, ...); pivot to wide.
    ent_key = "entity_id" if recs and "entity_id" in recs[0] else "entity"
    by_entity: dict[str, dict] = {}
    for r in recs:
        row = by_entity.setdefault(r.get(ent_key, "?"), {})
        field = r.get("field")
        if field:
            row[field] = r.get("value")
            src = r.get("source")
            if src:
                row.setdefault("sources", "")
                row["sources"] = (row["sources"] + " ; " + src).strip(" ;")
    headers = [c for c in COLUMNS if any(c in row for row in by_entity.values())] or COLUMNS
    return headers, list(by_entity.values())


# Columns derived from record metadata rather than a same-named field.
_DERIVED_COLS = ("entity_type", "confidence_entity_match", "data_quality",
                 "financial_sources", "other_sources")


def _field_to_col(cols: list[str]) -> dict:
    m = {c: c for c in cols if c not in _DERIVED_COLS}
    # legacy single-revenue records → total_revenue_2025 so old runs still project
    m["revenue"] = "total_revenue_2025"
    return m


# Columns whose values get canonicalized on export (raw stays in the JSON record).
_MONEY_COLS = set(FINANCIAL_COLUMNS) - _PERCENT_COLS - {"product_revenue_source"}


def _normalize_cell(col: str, v: str) -> str:
    from . import gate
    if not v:
        return v
    if col in _MONEY_COLS:
        return gate.normalize_money(v) or v
    if col in _PERCENT_COLS:
        return gate.normalize_percent(v) or v
    if col == "headcount":
        return gate.normalize_headcount(v) or v
    return v


def _record_to_row(rec: dict, cols: list[str] | None = None) -> dict:
    """Project one verifier record (verifier.md output) into a deliverable row."""
    from . import gate
    fields = rec.get("fields", {}) or {}
    em = rec.get("entity_match", {}) or {}
    cols = cols or get_columns()
    field_to_col = _field_to_col(cols)
    row = {c: "" for c in cols}

    fin_sources, other_sources = [], []
    for fname, f in fields.items():
        if not isinstance(f, dict):
            continue
        col = field_to_col.get(fname)
        if col and not row.get(col):
            row[col] = _normalize_cell(col, f.get("value", "") or "")
        src = f.get("source")
        if src:
            (fin_sources if fname in _FINANCIAL_SOURCE_FIELDS
             else other_sources).append(str(src))

    et = em.get("entity_type") or em.get("type") or ""
    if not et and isinstance(fields.get("entity_type"), dict):
        et = fields["entity_type"].get("value") or ""
    row["entity_type"] = gate.normalize_entity_type(et) or et
    row["confidence_entity_match"] = em.get("confidence", "") or ""
    if not row["brand_name"]:
        row["brand_name"] = rec.get("entity", "") or ""

    # de-dup each source list, preserve order (conflicts/flags/assumptions stay in
    # the JSON and surface in analysis.md — the deliverable table has no notes column)
    def _uniq(items: list[str]) -> str:
        seen, out = set(), []
        for s in items:
            if s not in seen:
                seen.add(s)
                out.append(s)
        return " ; ".join(out)

    row["financial_sources"] = _uniq(fin_sources)
    row["other_sources"] = _uniq(other_sources)

    # `unresolved: <field> — …` review_flags mark cells the pipeline could not
    # source (e.g. search quota exhausted): the cell stays EMPTY and gets a
    # yellow highlight in write_xlsx — never marker text in the data itself
    unresolved = []
    for fl in rec.get("review_flags") or []:
        s = str(fl)
        if s.lower().startswith("unresolved:"):
            fname = s.split(":", 1)[1].split("—")[0].strip()
            col = field_to_col.get(fname)
            if col:
                unresolved.append(col)
    row["_unresolved"] = unresolved   # meta key: not a column, csv ignores it

    # record + field quality states (additive layer over the gate verdict);
    # brand_name is derived from `entity` on export, so its absence as a
    # record field is not a gap
    q = gate.record_quality(
        rec, [f for f in field_to_col
              if f not in ("revenue", "brand_name") or f in fields])
    row["data_quality"] = f"{q['status']} · {q['coverage_pct']}%" + (
        f" · mandatory gaps: {', '.join(q['mandatory_gaps'][:3])}"
        if q["mandatory_gaps"] else "") + (
        f" · product-rev: {q['product_basis']}" if q.get("product_basis") else "")
    row["_low_conf"] = [field_to_col[f] for f in q["low_confidence"]
                        if f in field_to_col]      # peach fill per CELL
    return row


def read_records(path: Path | list[Path]) -> tuple[list[str], list[dict]]:
    """Read verifier record JSONs (a dir of *_record.json, a single file, or an
    explicit list of files — e.g. only the gate-accepted ones)."""
    if isinstance(path, (list, tuple)):
        files = [Path(p) for p in path]
        if not files:
            raise SystemExit("No accepted records to export.")
    elif path.is_dir():
        files = sorted(path.glob("*_record.json")) or sorted(path.glob("*.record.json"))
        if not files:
            raise SystemExit(f"No *_record.json files in {path}. Save each verifier "
                             f"result there, or use --csv.")
    else:
        files = [path]
    cols = get_columns()
    rows = []
    for fp in files:
        rec = json.loads(fp.read_text(encoding="utf-8"))
        rows.append(_record_to_row(rec, cols))
    return cols, rows


def write_csv(headers: list[str], rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in headers})
    print(f"[csv] wrote {len(rows)} rows → {path} (Airtable-ready)")


_RESP_WIDTHS = {"target": 22, "name": 22, "role": 30, "org": 24,
                "why_relevant": 46, "priority": 8, "telegram": 20, "phone": 18,
                "email": 26, "linkedin": 30, "profile_url": 34, "sources": 40,
                "confidence": 12, "verified_on": 14}


def write_respondents_sheet(path: Path, headers: list[str], rows: list[dict]) -> None:
    """Add or refresh a «Respondents» sheet. If the workbook already exists
    (the quantitative deliverable), the sheet joins it; otherwise a new
    workbook is created with «Respondents» as its first sheet (manual-only
    flow). Existing sheets are preserved; a stale «Respondents» is replaced."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        if "Respondents" in wb.sheetnames:
            del wb["Respondents"]
        ws = wb.create_sheet("Respondents", 0 if not wb.sheetnames else None)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Respondents"

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    ws.append(headers)
    for ci, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = _RESP_WIDTHS.get(name, 20)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        for ci in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=ci).alignment = wrap_top
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(path)
    print(f"[xlsx] Respondents sheet: {len(rows)} rows → {path}")


def write_xlsx(headers: list[str], rows: list[dict], out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Research"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    low_fill = PatternFill("solid", fgColor="FCE4D6")   # low-confidence highlight
    unresolved_fill = PatternFill("solid", fgColor="FFF599")   # yellow: needs a new source
    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws.append(headers)
    for ci, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = _WIDTHS.get(name, 22)

    conf_cols = [h for h in headers if "confidence" in h.lower()]
    hidx = {h: i + 1 for i, h in enumerate(headers)}
    for r in rows:
        values = [r.get(h, "") for h in headers]
        ws.append(values)
        ri = ws.max_row
        low = any(str(r.get(cc, "")).strip().lower() == "low" for cc in conf_cols)
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = wrap_top
            if low:
                cell.fill = low_fill
        for col in r.get("_low_conf", []):
            if col in hidx:
                ws.cell(row=ri, column=hidx[col]).fill = low_fill
        for col in r.get("_unresolved", []):
            if col in hidx:
                ws.cell(row=ri, column=hidx[col]).fill = unresolved_fill

    # Two-level drill-down on the financial block: outline level 1 = the whole
    # block (product_revenue_source is its level-1 summary column), level 2 =
    # each year's columns. Excel's 1/2/3 buttons collapse: [1] no financials,
    # [2] year groups visible, [3] everything.
    year_cols = set(_YEAR_COLS)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    for name in FINANCIAL_COLUMNS:
        if name in col_idx:
            dim = ws.column_dimensions[get_column_letter(col_idx[name])]
            dim.outline_level = 2 if name in year_cols else 1
    if any(n in col_idx for n in FINANCIAL_COLUMNS):
        ws.sheet_properties.outlinePr.summaryRight = True

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[xlsx] wrote {len(rows)} rows × {len(headers)} cols → {out}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Export the research table to Excel / Airtable-ready CSV.")
    src = ap.add_mutually_exclusive_group()
    src.add_argument("--from-records", type=Path,
                     help="dir (or file) of verifier *_record.json to project into the table")
    src.add_argument("--csv", type=Path, help="research CSV to render")
    src.add_argument("--from-kb", action="store_true", help="read db/kb.sqlite consolidated_record")
    src.add_argument("--template", type=Path,
                     help="write a blank research CSV (seeded from inputs/input_entities.csv) and exit")
    ap.add_argument("--subsector", default=None,
                    help="with --template: only seed rows whose 'subsector' matches (e.g. 'IaaS/GPU')")
    ap.add_argument("--out", type=Path, default=ROOT / "outputs" / "research_table.xlsx",
                    help="output .xlsx path")
    ap.add_argument("--csv-out", type=Path, default=ROOT / "outputs" / "research_table.csv",
                    help="Airtable-ready CSV to write when building from records/KB")
    ap.add_argument("--db", type=Path, default=ROOT / "db" / "kb.sqlite")
    args = ap.parse_args()

    if args.template:
        write_template(args.template, args.subsector)
        return

    if args.from_records:
        headers, rows = read_records(args.from_records)
        write_csv(headers, rows, args.csv_out)   # projection → CSV (Airtable) + xlsx
    elif args.from_kb:
        headers, rows = read_kb(args.db)
        write_csv(headers, rows, args.csv_out)
    elif args.csv:
        headers, rows = read_csv(args.csv)
    else:
        ap.error("give one of --from-records, --csv, --from-kb, or --template")

    write_xlsx(headers, rows, args.out)


if __name__ == "__main__":
    main()
